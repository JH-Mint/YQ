#coding=utf-8
#读取excel表格的模块
import openpyxl

from api_auto.api_auto.uitls.handle_path import *

#定义一个基类
class Handle_Excel(object):
    """
    封装一个读取excel表格的工具类
    """
    def __init__(self,filename,sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name

    def open(self):
        """
        封装一个打开excel表格工具方法
        :return:
        """
        #通过调用openpyxl模块下的load_workbook函数读取filename文件的内容
        self.wb = openpyxl.load_workbook(self.filename)
        print(self.wb)
        #通过self.wb这个excel文件对象读取对应的sheet页面信息
        self.sh = self.wb[self.sheet_name]
        print(self.sh)

    def read_data(self):
        """
        封装一个读取excel表格的工具方法
        :return:
        """
        self.open()  #调用open实例方法，获取sheet的页面信息
        datas = list(self.sh.rows)   #把每一行的数据信息放到列表中
        # for i in datas[0]:
        #     print(i.value)
            #列表解析式
        title =[i.value for i in datas[0]]
        print(title)
        #创建一个空的列表，用来接收所有的测试用例
        cases = []
        for j in datas[1:]: #对第2/3/4/5行进行for循环遍历
            values = [k.value for k in j]       #通过对j进行遍历，用k.value获取每个用例对象当中value值，然后放到列表当中
            case = dict(zip(title,values))  #把title列表和values用例一一对应打包，放入到一个字典当中
            cases.append(case) #把每个用例追加到列表当中
        print(cases)
        return cases

    def write_excel(self,row,column,value):
        """
        封装一个往eccel表中写入测试结果的工具方法
        :return:
        """
        self.open()
        #往固定的row行和column列写入value值
        self.sh.cell(row,column,value)
        #保存excel表格，并进行关闭
        self.wb.save(self.filename)





if __name__ == '__main__':

    path = os.path.join(data_path,"apicase.xlsx")
    h = Handle_Excel(path,"login")

    # h.write_excel(6,1,"麦钊鹏")



































































